''' '''
import configparser
import os
import time
import sys

import numpy as np
import astropy.units as u
from astropy.table import Table
from openpyxl import load_workbook

from marxs.missions.arcus.defaults import DefaultSource, DefaultPointing
from marxs.missions.arcus.generate_rowland import make_rowland_from_d_BF_R_f
from marxs.missions.arcus.arcus import defaultconf
from marxs.missions.arcus.ralfgrating import RegularGrid
from marxs.missions.arcus.spo import spos_large, spos_large_pos4d
from marxs.missions.arcus.arcus import PerfectArcus
from marxs.missions.mitsnl.catgrating import InterpolateEfficiencyTable


cfgpath = [os.path.join(os.path.dirname(sys.modules[__name__].__file__), '..', 'site.cfg')]
'Path list to search for configuration files.'

def get_path(name):
    '''Get path name info from site.cfg file in root directory.

    If a path does not exist, it will be created.
    '''
    conf = configparser.ConfigParser()
    cfgfile = conf.read(cfgpath)
    if not cfgfile:
        raise Exception("No config file with path specifications found. File must be called 'site.py' and be located in one of the following directories: {}".format(cfgpath))
    path = conf.get("Path", name)
    if not os.path.exists(path):
        os.makedirs(path)
    return path


conf = defaultconf.copy()
conf['spo_geom'] = spos_large
conf['spo_pos4d'] = spos_large_pos4d

class BigSPOArcus(PerfectArcus):
    gratings_class = RegularGrid


def sheet_to_table(sheet):
    # construct column names
    colnames = [sheet['B8'].value, sheet['C8'].value] + [cell.value for cell in sheet[9][3:]]
    # read data
    tab = Table([a for a in sheet.iter_cols(min_row=10, min_col=2, values_only=True)],
                names=colnames)
    # add units
    tab['wavelength'].unit = u.Unit(sheet['B9'].value.replace('[', '').replace(']', ''))
    tab['angle'].unit = u.Unit(sheet['C9'].value.replace('[', '').replace(']', ''))
    # Read meta data, which is saved at fixed location
    for i in [2,3,4,5]:
        tab.meta[sheet.cell(row=i, column=2).value] = sheet.cell(row=i, column=3).value * u.Unit(sheet.cell(row=i, column=4).value)
    tab.meta[sheet.cell(row=6, column=2).value] = sheet.cell(row=6, column=3).value

    # Drop the last three columns which have formulas in them
    tab.keep_columns(tab.colnames[:-3])
    return tab

efficiency = {}
wb = load_workbook('../inputdata/RCWA_for_grating_parameters.xlsx')
for sheet in wb.sheetnames[1:]:
    efficiency[sheet] = sheet_to_table(wb[sheet])
    efficiency[sheet].meta['sheetname'] = sheet

n_photons = 20000
wave = np.arange(8., 40.1, 1.) * u.Angstrom
#wave = [10., 15., 20., 25., 30., 35., 37.5] * u.Angstrom
energies = wave.to(u.keV, equivalencies=u.spectral()).value

mypointing = DefaultPointing()
mysource = DefaultSource()

# Make input photon list with grid of discrete energies
photons_in = mysource.generate_photons(n_photons * len(wave) * u.s)
photons_in = mypointing(photons_in)

for i, e in enumerate(energies):
    photons_in['energy'][i * n_photons: (i + 1) * n_photons] = e


arr_R = np.arange(5900., 6001., 100.)
arr_d_BF = np.arange(550., 751., 50.)
arr_blaze = np.arange(1.2, 2.21, 0.2)

def iterate_dBFblaze(photons_in, arr_d_BF, arr_blaze, conf, R, f=11880., metainfo={}, prefix=''):
    for d_BF in arr_d_BF:
        conf = conf | make_rowland_from_d_BF_R_f(d_BF, R, f)
        for blaze in arr_blaze:
            conf['blazeang'] = blaze
            filename = f'{prefix}{d_BF:04.0f}_{blaze:04.1f}_{R:05.0f}.fits'
            filepath = os.path.join(get_path('grid2designtorus'), filename)
            # This is a trick for parallelization.
            # Write file if working on it - and if file exist skip.
            # So, can start several workers in independent terminals.
            if os.path.isfile(filepath):
                continue
            else:
                photons_in.write(filepath)
                print('d_BF: {:4.0f} mm - blaze: {:4.1f} deg - R: {:5.0f} mm - {}'.format(d_BF, blaze, R, time.ctime()))
                instrum = BigSPOArcus(channels=['1'], conf=conf)
                #tiltCATGratings(instrum, arcus.arcus.defaultconf['blazeang'], blaze)
                photons_out = instrum(photons_in.copy())
                photons_out.meta['TORUS_R'] = R
                photons_out.meta['CIRCLE_R'] = conf['rowland_1'].r
                photons_out.meta['MAX_F'] = conf['f']
                photons_out.meta['BLAZE'] = blaze
                photons_out.meta['D_CHAN'] = d_BF
                photons_out.meta['N_PHOT'] = n_photons
                photons_out.meta['A_GEOM'] = instrum.elements[0].area.to(u.cm**2).value
                photons_out.meta.update(metainfo)
                # Keep only those columns absolutely needed for analysis
                # to reduce file size
                photons_out.keep_columns(['energy', 'probability', 'order',
                                          'circ_phi'])
                photons_out.write(filepath, overwrite=True)


for sheet in efficiency.keys():
    conf['order_selector'] = InterpolateEfficiencyTable(efficiency[sheet])
    # Currently, marx does not have tables for transmission through materials
    # other than Si and for coated gratings more complex stuff will happen anyway.
    # So, just treat Si part of bar here. L1 is as deep as the grating bars itself.
    # Won't make a difference fir these simulations, but for full consistency
    # does not hurt to set it here.
    bardepth, barwidth, out = sheet.split('-')
    conf['gratinggrid']['elem_args']['l1_dims']['bardepth'] = bardepth * u.micrometer
    iterate_dBFblaze(photons_in, arr_d_BF, arr_blaze, conf, R=5900, f=11880.,
                     prefix=sheet, metainfo={'grattype': sheet})

